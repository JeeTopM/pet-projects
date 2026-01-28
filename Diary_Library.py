import datetime
import os
import re
import sys
import threading
import tkinter as tk
import traceback
from pathlib import Path
from tkinter import filedialog, messagebox, scrolledtext, ttk
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook

# ======================
# === ВСПОМОГАТЕЛЬНЫЕ ===
# ======================

def to_number(value) -> int:
    """Безопасно преобразует значение в число."""
    if pd.isna(value):
        return 0
    try:
        return int(float(str(value)))
    except (ValueError, TypeError):
        return 0


def read_excel(file_path: Path):
    """Читает активный лист Excel."""
    if not file_path.exists():
        raise FileNotFoundError(f"Файл не найден: {file_path}")

    workbook = load_workbook(file_path)
    return workbook.active


def find_header(ws, keyword: str):
    """Ищет строку, содержащую указанный keyword."""
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        for cell_value in row:
            if cell_value and keyword in str(cell_value):
                return row_idx
    return None


def extract_table(ws, start_row: int):
    """Собирает данные из листа начиная со строки start_row."""
    data_rows = []
    for row_idx in range(start_row, ws.max_row + 1):
        row_data = [ws.cell(row=row_idx, column=col).value
                    for col in range(1, ws.max_column + 1)]

        if not any(cell is not None for cell in row_data):
            break

        data_rows.append(row_data)
    return data_rows


def parse_date(value):
    """Пытается разобрать дату из разных форматов."""
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time())
    str_val = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%Y.%m.%d"):
        try:
            return datetime.datetime.strptime(str_val, fmt)
        except ValueError:
            continue
    return None


def save_report(df: pd.DataFrame, source_path: Path, suffix: str) -> Path:
    """Сохраняет итоговый DataFrame в Excel рядом с исходным файлом."""
    new_path = source_path.parent / f"{source_path.stem}-{suffix}{source_path.suffix}"
    df.to_excel(new_path, index=False)
    return new_path


def format_month_name(date: datetime.datetime) -> str:
    """Форматирует название месяца на русском."""
    month_names = {
        1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель",
        5: "Май", 6: "Июнь", 7: "Июль", 8: "Август",
        9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь"
    }
    return f"{month_names[date.month]} {date.year}"


def create_monthly_report(data: List[Dict], week_col: str = "№ недели") -> pd.DataFrame:
    """
    Создает отчет с группировкой по месяцам.
    Агрегирует данные по неделям в пределах каждого месяца.
    """
    if not data:
        return pd.DataFrame()

    # Преобразуем данные в DataFrame
    df = pd.DataFrame(data)

    # Добавляем столбцы для месяца и года
    df['date'] = pd.to_datetime(df.get('date', pd.NaT))
    df['month_year'] = df['date'].apply(lambda x: format_month_name(x) if pd.notna(x) else '')
    df['month_num'] = df['date'].dt.month
    df['year'] = df['date'].dt.year
    df['week_num'] = df[week_col]  # Сохраняем номер недели

    # Группируем по месяцу и неделе, суммируя все числовые колонки
    numeric_cols = [col for col in df.columns if col not in ['date', 'month_year', 'month_num', 'year', week_col, 'week_num']]

    # Суммируем данные по неделям в пределах каждого месяца
    grouped = df.groupby(['year', 'month_num', 'month_year', 'week_num'])[numeric_cols].sum().reset_index()

    # Сортируем по дате
    grouped = grouped.sort_values(['year', 'month_num', 'week_num'])

    result_rows = []
    month_totals = {}
    current_month = None

    for _, row in grouped.iterrows():
        month_name = row['month_year']

        # Если начался новый месяц, добавляем заголовок месяца
        if month_name != current_month:
            if current_month is not None and month_totals:
                # Добавляем строку ИТОГО за предыдущий месяц
                total_row = {week_col: "ИТОГО"}
                for col in month_totals.keys():
                    total_row[col] = month_totals[col]
                result_rows.append(total_row)
                result_rows.append({})  # Пустая строка для разделения

            # Добавляем заголовок нового месяца
            result_rows.append({week_col: month_name})
            current_month = month_name
            month_totals = {col: 0 for col in numeric_cols}

        # Добавляем строку с агрегированными данными недели
        week_row = {week_col: f"Неделя {int(row['week_num'])}"}
        for col in numeric_cols:
            value = to_number(row[col])
            week_row[col] = value
            month_totals[col] += value
        result_rows.append(week_row)

    # Добавляем ИТОГО за последний месяц
    if month_totals:
        total_row = {week_col: "ИТОГО"}
        for col in month_totals.keys():
            total_row[col] = month_totals[col]
        result_rows.append(total_row)

    # Добавляем общее ВСЕГО
    result_rows.append({})  # Пустая строка для разделения

    # Вычисляем общие итоги по всем данным (не агрегированным)
    grand_total_row = {week_col: "ВСЕГО"}
    for col in numeric_cols:
        grand_total_row[col] = df[col].sum()
    result_rows.append(grand_total_row)

    return pd.DataFrame(result_rows)


# ======================
# === ОТЧЕТ 1. ПОЛЬЗОВАТЕЛИ ===
# ======================

def process_report_1(file_path: Path) -> Path | None:
    """1. Дневник библиотеки. Часть 1.1 — Пользователи."""
    ws = read_excel(file_path)
    header_row_idx = find_header(ws, "Дата")

    if not header_row_idx:
        raise ValueError("Не найден заголовок 'Дата'")

    data_rows = extract_table(ws, header_row_idx)

    header_row = next((i for i, row in enumerate(data_rows)
                       if len(row) > 2 and row[1] == 'Дата' and row[2] == 'Всего читателей'), None)
    if header_row is None:
        raise ValueError("Не найдена строка с заголовками данных!")

    df = pd.DataFrame(data_rows[header_row + 1:], columns=data_rows[header_row])

    temp_data = []
    for _, row in df.iterrows():
        date_val = parse_date(row.iloc[1])
        if date_val is None:
            continue

        week = date_val.isocalendar()[1]
        temp_data.append({
            "date": date_val,
            "№ недели": week,
            "0-6": to_number(row.iloc[7]),
            "7-9": to_number(row.iloc[8]),
            "10-14": to_number(row.iloc[9]),
            "15-17": to_number(row.iloc[10]),
            "18-35": to_number(row.iloc[11]),
            "36-55": to_number(row.iloc[13]),
            "56 и старше": to_number(row.iloc[14]),
        })

    if not temp_data:
        raise ValueError("Нет данных для обработки.")

    grouped = create_monthly_report(temp_data)
    return save_report(grouped, file_path, "пользователи")


# ======================
# === ОТЧЕТ 2. ЗАПИСЬ ЧИТАТЕЛЕЙ ===
# ======================

def process_report_2(file_path: Path) -> Path | None:
    """2. Статистика записи читателей по округу/библиотеке."""
    ws = read_excel(file_path)
    header_row_idx = find_header(ws, "Пункт книговыдачи / период")

    if not header_row_idx:
        raise ValueError("Не найден заголовок 'Пункт книговыдачи / период'")

    temp_data = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx <= header_row_idx or not row or not row[1]:
            continue

        date = parse_date(row[1])
        if not date:
            continue

        week = date.isocalendar()[1]
        temp_data.append({
            "date": date,
            "№ недели": week,
            "Договоры": to_number(row[2]),
        })

    if not temp_data:
        raise ValueError("Нет данных для обработки.")

    grouped = create_monthly_report(temp_data)
    return save_report(grouped, file_path, "запись-читателей")


# ======================
# === ОТЧЕТ 3. ПОСЕЩЕНИЯ ===
# ======================

def process_report_3(file_path: Path) -> Path | None:
    """3. Дневник библиотеки. Часть 1.2 — Посещения."""
    ws = read_excel(file_path)
    header_row_idx = find_header(ws, "Дата")

    if not header_row_idx:
        raise ValueError("Не найден заголовок 'Дата'")

    data_rows = extract_table(ws, header_row_idx)
    data_start_row = next((i for i, row in enumerate(data_rows)
                           if len(row) > 1 and row[1] == "Дата"), None)
    if data_start_row is None:
        raise ValueError("Не найдена строка с заголовками данных.")

    df = pd.DataFrame(data_rows[data_start_row + 1:], columns=data_rows[data_start_row])
    temp_data = []

    for _, row in df.iterrows():
        date_val = parse_date(row.iloc[1])
        if not date_val:
            continue

        week = date_val.isocalendar()[1]
        temp_data.append({
            "date": date_val,
            "№ недели": week,
            "Посещения": to_number(row.iloc[4]) + to_number(row.iloc[7]) +
                         to_number(row.iloc[9]) + to_number(row.iloc[13]),
            "КДФ": to_number(row.iloc[12]),
            "Почта": to_number(row.iloc[21]),
            "Телефон": to_number(row.iloc[20]),
            "В стационарных условиях": to_number(row.iloc[16]),
            "Справки 1": to_number(row.iloc[17]),
            "Справки 2": to_number(row.iloc[18]),
            "Справки 3": to_number(row.iloc[19]),
        })

    if not temp_data:
        raise ValueError("Нет данных для обработки.")

    grouped = create_monthly_report(temp_data)
    return save_report(grouped, file_path, "посещения")


# ======================
# === ОТЧЕТ 4. КНИГОВЫДАЧА ===
# ======================

def process_report_4(file_path: Path) -> Path | None:
    """4. Дневник библиотеки — статистика книговыдачи."""
    ws = read_excel(file_path)
    header_row_idx = find_header(ws, "Пункт книговыдачи")

    if not header_row_idx:
        raise ValueError("Не найден заголовок 'Пункт книговыдачи / период'")

    # data_start = next((r for r, row in enumerate(ws.iter_rows(values_only=True), 1) if row[1] and isinstance(row[1], str) and "2026-" in row[1]), None)

    data_start = next((r for r, row in enumerate(ws.iter_rows(values_only=True), 1)
                    if row[1] and isinstance(row[1], str)
                    and re.search(r'\b\d{4}-', row[1])), None)


    if not data_start:
        raise ValueError("Не найдено начало таблицы с данными.")

    data_rows = extract_table(ws, data_start)
    temp_data = []

    for row in data_rows:
        if not row or not row[1]:
            continue

        date_val = parse_date(row[1])
        if not date_val:
            continue

        week = date_val.isocalendar()[1]
        children_1 = sum(to_number(row[i]) for i in [5, 6, 7])
        children_2 = sum(to_number(row[i]) for i in [8])
        youth = sum(to_number(row[i]) for i in [9])

        temp_data.append({
            "date": date_val,
            "№ недели": week,
            "Всего": to_number(row[2]),
            "Детям до 14 лет вкл.": children_1,
            "Подростки 15-17 лет": children_2,
            "Молодежь 18-35 лет": youth
        })

    if not temp_data:
        raise ValueError("Нет данных для обработки.")

    grouped = create_monthly_report(temp_data)
    return save_report(grouped, file_path, "книговыдача")


# ======================
# === GUI ПРИЛОЖЕНИЕ ===
# ======================

class LibraryReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Обработчик отчетов библиотеки")
        self.root.geometry("800x600")

        # Установка иконки (если есть)
        try:
            if getattr(sys, 'frozen', False):
                # Если запущен как EXE
                base_path = sys._MEIPASS
            else:
                # Если запущен как скрипт
                base_path = os.path.dirname(__file__)

            icon_path = os.path.join(base_path, "icon.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
        except:
            pass

        self.setup_ui()
        self.file_path = None

    def setup_ui(self):
        # Заголовок
        title_frame = ttk.Frame(self.root, padding="10")
        title_frame.pack(fill="x")

        title_label = ttk.Label(
            title_frame,
            text="📚 Обработчик отчетов библиотеки",
            font=("Arial", 16, "bold")
        )
        title_label.pack()

        subtitle_label = ttk.Label(
            title_frame,
            text="С группировкой данных по месяцам",
            font=("Arial", 10)
        )
        subtitle_label.pack()

        # Основной контейнер
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill="both", expand=True)

        # Выбор файла
        file_frame = ttk.LabelFrame(main_frame, text="Выберите файл отчета", padding="10")
        file_frame.pack(fill="x", pady=(0, 10))

        self.file_path_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_path_var, width=60)
        file_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        browse_btn = ttk.Button(
            file_frame,
            text="Обзор...",
            command=self.browse_file
        )
        browse_btn.pack(side="right")

        # Выбор типа отчета
        report_frame = ttk.LabelFrame(main_frame, text="Выберите тип отчета", padding="10")
        report_frame.pack(fill="x", pady=(0, 10))

        self.report_type = tk.IntVar(value=1)

        reports = [
            ("1. Дневник библиотеки. Часть 1.1 – Пользователи", 1),
            ("2. Статистика записи читателей по округу/библиотеке", 2),
            ("3. Дневник библиотеки. Часть 1.2 – Посещения", 3),
            ("4. Дневник библиотеки – статистика книговыдачи", 4)
        ]

        for text, value in reports:
            radio = ttk.Radiobutton(
                report_frame,
                text=text,
                variable=self.report_type,
                value=value
            )
            radio.pack(anchor="w", pady=2)

        # Описание формата
        desc_frame = ttk.LabelFrame(main_frame, text="Формат выходного отчета", padding="10")
        desc_frame.pack(fill="x", pady=(0, 10))

        desc_text = """
        • Название месяца
        • Данные по неделям (агрегированные)
        • ИТОГО за месяц
        • Пустая строка
        • Следующий месяц...
        • ВСЕГО (общий итог)
        """

        desc_label = ttk.Label(desc_frame, text=desc_text, justify="left")
        desc_label.pack(anchor="w")

        # Кнопка обработки
        button_frame = ttk.Frame(main_frame)
        button_frame.pack(fill="x", pady=(10, 0))

        self.process_btn = ttk.Button(
            button_frame,
            text="Обработать отчет",
            command=self.process_report,
            style="Accent.TButton"
        )
        self.process_btn.pack(side="left", padx=(0, 10))

        self.open_folder_btn = ttk.Button(
            button_frame,
            text="Открыть папку с файлами",
            command=self.open_folder,
            state="disabled"
        )
        self.open_folder_btn.pack(side="left")

        # Лог сообщений
        log_frame = ttk.LabelFrame(main_frame, text="Лог обработки", padding="10")
        log_frame.pack(fill="both", expand=True, pady=(10, 0))

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=10,
            wrap=tk.WORD,
            font=("Courier New", 9)
        )
        self.log_text.pack(fill="both", expand=True)

        # Статус бар
        self.status_var = tk.StringVar(value="Готов к работе")
        status_bar = ttk.Label(
            self.root,
            textvariable=self.status_var,
            relief=tk.SUNKEN,
            anchor=tk.W,
            padding=(10, 5)
        )
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        # Стили
        style = ttk.Style()
        style.configure("Accent.TButton", font=("Arial", 10, "bold"))

    def browse_file(self):
        filetypes = (
            ("Excel files", "*.xlsx *.xls"),
            ("All files", "*.*")
        )

        filename = filedialog.askopenfilename(
            title="Выберите файл отчета",
            filetypes=filetypes
        )

        if filename:
            self.file_path_var.set(filename)
            self.file_path = Path(filename)
            self.log_message(f"Выбран файл: {filename}")

    def log_message(self, message: str):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def process_report(self):
        if not self.file_path_var.get():
            messagebox.showerror("Ошибка", "Пожалуйста, выберите файл отчета")
            return

        if not self.file_path.exists():
            messagebox.showerror("Ошибка", f"Файл не найден:\n{self.file_path}")
            return

        report_num = self.report_type.get()

        processors = {
            1: ("пользователи", process_report_1),
            2: ("запись-читателей", process_report_2),
            3: ("посещения", process_report_3),
            4: ("книговыдача", process_report_4),
        }

        report_name, processor = processors[report_num]

        # Отключаем кнопку на время обработки
        self.process_btn.config(state="disabled")
        self.open_folder_btn.config(state="disabled")
        self.status_var.set("Обработка...")

        # Запускаем в отдельном потоке
        thread = threading.Thread(
            target=self.run_processor,
            args=(processor, report_name),
            daemon=True
        )
        thread.start()

    def run_processor(self, processor, report_name: str):
        try:
            self.log_message(f"Начинаю обработку отчета '{report_name}'...")

            # Выполняем обработку
            result_path = processor(self.file_path)

            # Обновляем GUI в основном потоке
            self.root.after(0, self.on_processing_complete, result_path, report_name)

        except Exception as e:
            error_msg = f"Ошибка обработки: {str(e)}\n{traceback.format_exc()}"
            self.root.after(0, self.on_processing_error, error_msg)

    def on_processing_complete(self, result_path: Path, report_name: str):
        self.process_btn.config(state="normal")
        self.open_folder_btn.config(state="normal")
        self.status_var.set("Обработка завершена")

        self.log_message(f"✅ Отчет успешно сохранен!")
        self.log_message(f"📁 Файл: {result_path}")

        messagebox.showinfo(
            "Успешно!",
            f"Отчет '{report_name}' успешно обработан!\n\n"
            f"Файл сохранен как:\n{result_path.name}"
        )

    def on_processing_error(self, error_msg: str):
        self.process_btn.config(state="normal")
        self.status_var.set("Ошибка обработки")

        self.log_message(f"❌ Ошибка при обработке:")
        self.log_message(error_msg)

        messagebox.showerror(
            "Ошибка обработки",
            "Произошла ошибка при обработке файла.\n"
            "Подробности смотрите в логе."
        )

    def open_folder(self):
        if self.file_path and self.file_path.exists():
            import subprocess
            folder_path = str(self.file_path.parent)

            try:
                if sys.platform == "win32":
                    os.startfile(folder_path)
                elif sys.platform == "darwin":  # macOS
                    subprocess.Popen(["open", folder_path])
                else:  # Linux
                    subprocess.Popen(["xdg-open", folder_path])
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть папку: {e}")
        else:
            messagebox.showwarning("Внимание", "Сначала выберите файл")


# ======================
# === ТОЧКА ВХОДА ===
# ======================

def main():
    root = tk.Tk()
    app = LibraryReportApp(root)

    # Центрирование окна
    root.update_idletasks()
    width = root.winfo_width()
    height = root.winfo_height()
    x = (root.winfo_screenwidth() // 2) - (width // 2)
    y = (root.winfo_screenheight() // 2) - (height // 2)
    root.geometry(f"{width}x{height}+{x}+{y}")

    root.mainloop()


if __name__ == "__main__":
    main()